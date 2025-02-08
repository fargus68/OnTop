from openpyxl.reader.excel import load_workbook

def read_password():
    wb = load_workbook(filename = "c:/temp/TTPlaner_Password.xlsx")
    ws = wb.worksheets[0]
    cell = ws.cell(row=1, column=1)
    return cell.value
