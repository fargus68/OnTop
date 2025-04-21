from openpyxl.reader.excel import load_workbook

def read_records_from_excel(filename, record, header_rows = 1, multiple_records_possible = False):
    #file_full_path = "./Data/Flows/" + filename + ".xlsx"
    wb = load_workbook(filename = filename)
    print(wb.sheetnames)
    ws = wb.worksheets[0]

    all_rows = []
    row_ix = 1
    record_found = False

    for row in ws:
        #print(row)
        current_row = []
        col_ix = 1
        #record_found = False

        if row_ix <= header_rows:
            for header_cell in row:
                current_row.append(header_cell.value)
            all_rows.append(current_row)
        else:
            for cell in row:
                if col_ix == 1:
                    if cell.value != record:
                        break
                    else:
                        record_found = True
                current_row.append(cell.value)
                col_ix += 1

            if record_found:
                all_rows.append(current_row)
                if not multiple_records_possible:
                    break

        print(current_row)
        row_ix += 1

    if not record_found:
        raise Exception("Record not found!")

    return all_rows

def read_structural_record_from_excel(filename, record):
    #file_full_path = "./Data/Flows/" + filename + ".xlsx"
    wb = load_workbook(filename = filename)
    print(wb.sheetnames)
    ws = wb.worksheets[0]

    all_rows = []
    row_ix = 1
    record_found = False

    for row in ws:
        #print(row)
        current_row = []
        col_ix = 1
        #record_found = False

        #if row_ix <= header_rows:
        if row_ix <= 1:
            for header_cell in row:
                current_row.append(header_cell.value)
            all_rows.append(current_row)
        else:
            for cell in row:
                if col_ix == 1:
                    if cell.value != record:
                        break
                    else:
                        record_found = True
                current_row.append(cell.value)
                col_ix += 1

            if record_found:
                all_rows.append(current_row)
                break

        print(current_row)
        row_ix += 1

    if not record_found:
        raise Exception("Record not found!")

    return all_rows

def read_technical_records_from_excel(filename, record, aut = "Chromium", multiple_records_possible = False):
    #file_full_path = "./Data/Flows/" + filename + ".xlsx"
    wb = load_workbook(filename = filename)
    print(wb.sheetnames)
    ws = wb.worksheets[0]

    all_rows = []
    row_ix = 1
    aut_found = False
    control_header_found = False
    record_found = False

    for row in ws:
        print(row)
        current_row = []
        col_ix = 1
        relevant_row = False

        for cell in row:
            if col_ix == 1:
                if cell.value != aut and cell.value != record and cell.value != "Record / Control":
                    break
                else:
                    relevant_row = True
                    if cell.value == aut:
                        aut_found = True
                    if cell.value == "Record / Control":
                        control_header_found = True
                    if cell.value == record:
                        record_found = True
            current_row.append(cell.value)
            col_ix += 1

        if relevant_row:
            all_rows.append(current_row)
            if record_found and not multiple_records_possible:
                break
        else:
            if record_found and multiple_records_possible:
                break

        print(current_row)
        row_ix += 1

    if not aut_found:
        raise Exception("AUT row not found!")

    if not control_header_found:
        raise Exception("Control header row not found!")

    if not record_found:
        raise Exception("Record row not found!")

    return all_rows